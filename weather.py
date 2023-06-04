import os
import re
import openpyxl
import logging
import chardet
from concurrent.futures import ThreadPoolExecutor
from PyPDF2 import PdfReader

# 设置日志记录器
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.FileHandler('weather_extraction.log')
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

# 定义文件路径和正则表达式
folder_path = r'C:\Users\Alize\Desktop\weather in Barranquilla'
file_regex = re.compile(r'^BOLETIN INFORMATIVO.*\.pdf$')

# 定义Excel文件路径和工作表名称
excel_file = r'C:\Users\Alize\Desktop\weather in Barranquilla\Weather in Barranquilla.xlsx'
sheet_name = 'Weather'

# 定义批量处理大小和线程池大小
batch_size = 10
max_workers = 4

# 创建Excel文件和工作表对象
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = sheet_name

# 添加Excel表头信息
headers = ['天气', '风向', '风速（节）', '环境温度（℃）', '海浪高度（m）', '航行建议']
for i, header in enumerate(headers):
    worksheet.cell(row=1, column=i+2, value=header)

def extract_pdf_info(pdf_path):
    """提取PDF文件中的信息"""
    with open(pdf_path, 'rb') as pdf_file:
        # 检测编码并指定编码打开文件
        pdf_reader = PdfReader(pdf_file)
        page = pdf_reader.pages[0]
        text = page.extract_text()

        # 提取日期信息
        date_regex = re.compile(r'Pronóstico para el\s+(.*?)\n\n\s+Área Costera')
        date_match = date_regex.search(text)
        if date_match:
            date = date_match.group(1)

            # 提取天气信息
            weather_regex = re.compile(r'Tiempo:\s+(.*)\n')
            weather_match = weather_regex.search(text)
            if weather_match:
                weather = weather_match.group(1)

                # 提取其他信息
                info_regex = re.compile(r'Vientos:\s+(.*)\n\s+Temperatura')
                info_match = info_regex.search(text)
                if info_match:
                    info = info_match.group(1).split(',')

                    # 提取风向、风速、环境温度和海浪高度信息
                    wind_direction = info[0].strip()
                    wind_speed = info[1].strip()
                    temperature = info[2].strip()
                    wave_height = info[3].strip()

                    # 提取航行建议信息
                    advice_regex = re.compile(r'Recomendación:\s+(.*)\n\n')
                    advice_match = advice_regex.search(text)
                    if advice_match:
                        advice = advice_match.group(1)

                        

                        # 将提取的信息写入Excel表格中
                        row = worksheet.max_row + 1
                        worksheet.cell(row=row, column=1, value=date)
                        worksheet.cell(row=row, column=2, value='Área Costera')
                        worksheet.cell(row=row, column=3, value=weather)
                        worksheet.cell(row=row, column=4, value=wind_direction)
                        worksheet.cell(row=row, column=5, value=wind_speed)
                        worksheet.cell(row=row, column=6, value=temperature)
                        worksheet.cell(row=row, column=7, value=wave_height)
if __name__ == '__main__':
    pdf_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if file_regex.match(f)]
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for i in range(0, len(pdf_files), batch_size):
            batch_files = pdf_files[i:i+batch_size]
            futures = [executor.submit(extract_pdf_info, pdf_path) for pdf_path in batch_files]
            for future in futures:
                try:
                    future.result()
                except Exception as e:
                    logger.error(f'Error extracting PDF info from {future} - {e}')
    workbook.save(excel_file)





